# pip install numpy pillow tensorflow openpyxl

import re
import shutil
import argparse
from pathlib import Path

import numpy as np
from PIL import Image, ImageOps
import tensorflow as tf

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# =====================================================
# CONFIG DEFAULT
# =====================================================
MODEL_PATH = "./0_model_training/model_unquant.tflite"
LABELS_PATH = "./0_model_training/labels.txt"
SRC_DIR = "./2_images"
DST_DIR = "./3_scan_output"
LOG_PATH = "./tm_scan_log.xlsx"  # FIX: .xlsx (bukan .xlxs)

NEG_THRESHOLD = 0.7   # NEG minimal untuk dianggap "NEG kuat"
KWH_THRESHOLD = 0.7   # KWH minimal untuk dianggap "TRUE"
IMG_EXT = [".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tif", ".tiff"]

# =====================================================
# ARGUMENT PARSER
# =====================================================
def parse_args():
    p = argparse.ArgumentParser("TFLite Scanner – Teachable Machine Compatible (Excel + gambar + status)")
    p.add_argument("--model", default=MODEL_PATH)
    p.add_argument("--labels", default=LABELS_PATH)
    p.add_argument("--src", default=SRC_DIR)
    p.add_argument("--dst", default=DST_DIR)
    p.add_argument("--log", default=LOG_PATH)

    p.add_argument("--neg_threshold", type=float, default=NEG_THRESHOLD)
    p.add_argument("--kwh_threshold", type=float, default=KWH_THRESHOLD)

    # aturan FALSE tambahan (opsional)
    p.add_argument("--expected_idpel_len", type=int, default=0,
                   help="Jika >0, idpel kurang 1 digit => FALSE (contoh expected 12, terbaca 11)")
    p.add_argument("--expected_stand_len", type=int, default=0,
                   help="Jika >0, stand kurang 1 digit => FALSE")

    # embed thumbnail gambar ke Excel (default ON)
    p.add_argument("--embed_images", dest="embed_images", action="store_true", default=True)
    p.add_argument("--no_embed_images", dest="embed_images", action="store_false")
    p.add_argument("--thumb_size", type=int, default=160, help="Ukuran max thumbnail (px)")

    return p.parse_args()

# =====================================================
# LOAD LABELS
# =====================================================
def load_labels(path):
    labels = [x.strip() for x in open(path, "r", encoding="utf-8") if x.strip()]
    if len(labels) != 2:
        raise ValueError("Model harus 2 kelas: KWH & NEG")

    kwh_idx = next(i for i, l in enumerate(labels) if "kwh" in l.lower())
    neg_idx = 1 - kwh_idx
    return labels, kwh_idx, neg_idx

# =====================================================
# LOAD MODEL
# =====================================================
def load_interpreter(model_path):
    itp = tf.lite.Interpreter(model_path=model_path)
    itp.allocate_tensors()

    in_det = itp.get_input_details()[0]
    out_det = itp.get_output_details()[0]

    h, w = in_det["shape"][1:3]
    dtype = in_det["dtype"]

    return itp, in_det, out_det, (h, w, dtype)

# =====================================================
# PREPROCESS (TEACHABLE MACHINE STYLE)
# =====================================================
def preprocess_image(img_path, size, dtype):
    img = Image.open(img_path).convert("RGB")
    img = ImageOps.fit(img, size, Image.Resampling.LANCZOS)

    arr = np.asarray(img).astype(np.float32)

    # normalize: [-1, 1]
    arr = (arr / 127.5) - 1.0

    if dtype == np.uint8:
        arr = ((arr + 1) * 127.5).clip(0, 255).astype(np.uint8)

    return np.expand_dims(arr, axis=0)

# =====================================================
# SOFTMAX SAFE
# =====================================================
def softmax(x):
    e = np.exp(x - np.max(x))
    return e / e.sum()

# =====================================================
# BEST-EFFORT EXTRACT IDPEL & STAND
# =====================================================
def extract_idpel_and_stand(img_path: Path):
    """
    Default rule:
    - Ambil grup digit terpanjang (>= 8 digit) sebagai idpel
    - Ambil grup digit berikutnya (>= 3 digit) sebagai stand (jika ada)

    Jika format nama file kamu beda, tinggal ubah fungsi ini.
    """
    text = f"{img_path.parent.name}_{img_path.stem}"
    groups = re.findall(r"\d+", text)
    if not groups:
        return "", ""

    # idpel: digit group terpanjang yang masuk akal
    idpel = ""
    for g in sorted(groups, key=len, reverse=True):
        if len(g) >= 8:
            idpel = g
            break
    if not idpel:
        idpel = groups[0]

    # stand: digit group pertama selain idpel
    stand = ""
    for g in groups:
        if g == idpel:
            continue
        if len(g) >= 3:
            stand = g
            break

    return idpel, stand

# =====================================================
# THUMBNAIL FOR EXCEL
# =====================================================
def make_thumbnail(src_path: Path, thumb_dir: Path, max_px: int) -> Path:
    thumb_dir.mkdir(parents=True, exist_ok=True)
    thumb_path = thumb_dir / f"{src_path.stem}__thumb.png"
    if thumb_path.exists():
        return thumb_path

    img = Image.open(src_path).convert("RGB")
    img.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)
    img.save(thumb_path, format="PNG", optimize=True)
    return thumb_path

# =====================================================
# BUILD EXCEL
# =====================================================
def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "scan_log"

    headers = [
        "idpel", "stand", "filename",
        "P(KWH)", "P(NEG)", "pred",
        "status", "reason",
        "image", "copied_to", "src_path"
    ]
    ws.append(headers)

    # header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # column widths
    widths = {
        "A": 16, "B": 10, "C": 32,
        "D": 10, "E": 10, "F": 8,
        "G": 10, "H": 40,
        "I": 18, "J": 40, "K": 60
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return wb, ws

# =====================================================
# MAIN
# =====================================================
def main():
    args = parse_args()

    labels, kwh_idx, neg_idx = load_labels(args.labels)
    itp, in_det, out_det, meta = load_interpreter(args.model)
    h, w, dtype = meta

    Path(args.dst).mkdir(parents=True, exist_ok=True)
    thumb_dir = Path(args.dst) / ".thumbs"

    wb, ws = build_workbook()

    total = copied = 0

    for img_path in Path(args.src).rglob("*"):
        if img_path.suffix.lower() not in IMG_EXT:
            continue

        try:
            # ===== inference =====
            inp = preprocess_image(img_path, (w, h), dtype)
            itp.set_tensor(in_det["index"], inp)
            itp.invoke()
            raw = np.squeeze(itp.get_tensor(out_det["index"]))

            # handle sigmoid / softmax output
            if raw.ndim == 0:
                probs = np.array([1 - raw, raw], dtype=np.float32)
            else:
                probs = softmax(raw.astype(np.float32))

            p_kwh = float(probs[kwh_idx])
            p_neg = float(probs[neg_idx])

            pred = "KWH" if p_kwh >= p_neg else "NEG"

            # ===== status rules =====
            # TRUE hanya kalau: pred=KWH dan confidence KWH >= kwh_threshold
            # FALSE kalau:
            # - pred=NEG (umumnya: pagar / tidak cocok / stand meter tidak terlihat)
            # - keyword 'pagar' di nama file/folder
            # - idpel/stand kurang 1 digit (opsional)
            status = True
            reasons = []

            idpel, stand = extract_idpel_and_stand(img_path)

            if pred == "NEG":
                status = False
                if p_neg >= args.neg_threshold:
                    reasons.append(f"NEG>={args.neg_threshold:.2f}")
                else:
                    reasons.append("pred=NEG")

            if pred == "KWH" and p_kwh < args.kwh_threshold:
                status = False
                reasons.append(f"KWH<{args.kwh_threshold:.2f}")

            # FALSE ketika ada gambar pagar (berdasarkan keyword)
            if "pagar" in img_path.name.lower() or "pagar" in str(img_path.parent).lower():
                status = False
                reasons.append("pagar")

            # FALSE ketika idpel kurang 1 digit
            if args.expected_idpel_len and idpel and len(idpel) == args.expected_idpel_len - 1:
                status = False
                reasons.append(f"idpel kurang 1 digit ({len(idpel)}/{args.expected_idpel_len})")

            # FALSE ketika stand kurang 1 digit
            if args.expected_stand_len and stand and len(stand) == args.expected_stand_len - 1:
                status = False
                reasons.append(f"stand kurang 1 digit ({len(stand)}/{args.expected_stand_len})")

            reason_text = "; ".join(dict.fromkeys(reasons))  # unique, keep order

            # ===== copy NEG kuat (sesuai script awal) =====
            copied_to = ""
            if pred == "NEG" and p_neg >= args.neg_threshold:
                dst = Path(args.dst) / img_path.name
                shutil.copy2(img_path, dst)
                copied_to = str(dst)
                copied += 1

            # ===== write excel row =====
            ws.append([
                idpel,
                stand,
                img_path.name,
                round(p_kwh, 4),
                round(p_neg, 4),
                pred,
                "TRUE" if status else "FALSE",
                reason_text,
                "",  # image placeholder
                copied_to,
                str(img_path),
            ])
            row = ws.max_row

            # cell alignment
            for c in range(1, 12):
                ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)

            # embed thumbnail image
            if args.embed_images:
                thumb_path = make_thumbnail(img_path, thumb_dir, args.thumb_size)
                xl_img = XLImage(str(thumb_path))
                xl_img.anchor = f"I{row}"  # column image
                ws.add_image(xl_img)
                ws.row_dimensions[row].height = 120

            total += 1
            if total % 25 == 0:
                print(f"[INFO] {total} image diproses...")

        except Exception as e:
            print(f"[ERROR] {img_path.name}: {e}")

    # highlight rows with FALSE
    false_fill = PatternFill("solid", fgColor="F8D7DA")  # light red
    for r in range(2, ws.max_row + 1):
        if ws[f"G{r}"].value == "FALSE":
            for col in range(1, 12):
                ws.cell(row=r, column=col).fill = false_fill

    Path(args.log).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.log)

    print("\n=== SELESAI ===")
    print(f"Total diproses : {total}")
    print(f"Dicopy (NEG)   : {copied}")
    print(f"Log (xlsx)     : {args.log}")

if __name__ == "__main__":
    main()