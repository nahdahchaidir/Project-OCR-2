#!/usr/bin/env python3
# ============================================================
# VERIFIKASI TANPA OCR:
# - TFLite Classifier (KWH/NEG) saja
# - Output: excel_idpel_kwh.xlsx dengan gambar thumbnail
#   Kolom: gambar | idpel  (HANYA 2 KOLOM)
#
# Install:
# pip install numpy pillow tensorflow openpyxl opencv-python
# ============================================================

import os
# (opsional) bantu kurangi noise TF saat run manual
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")
os.environ.setdefault("TF_ENABLE_ONEDNN_OPTS", "0")
os.environ.setdefault("CUDA_VISIBLE_DEVICES", "-1")

import re
import sys
import time
import shutil
import argparse
import json
import csv
from pathlib import Path
from datetime import datetime
import tempfile

import numpy as np
import cv2
import tensorflow as tf

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage


# ===================== DEFAULT CONFIG =====================
MODEL_PATH = "./0_model_training/model_unquant.tflite"
LABELS_PATH = "./0_model_training/labels.txt"
SRC_DIR = "./2_images"
DST_DIR = "./3_scan_output"
LOG_PATH = "./excel_idpel_kwh.xlsx"  # ✅ Format .xlsx untuk gambar

NEG_THRESHOLD = 0.70
KWH_THRESHOLD = 0.70
BLUR_THRESHOLD = 80.0
# Tambah threshold untuk deteksi foto gelap/tidak ada angka
BRIGHTNESS_THRESHOLD = 30.0  # Nilai rata-rata brightness minimum
CONTRAST_THRESHOLD = 20.0    # Nilai kontras minimum

EXPECTED_IDPEL_LEN = 12

IMG_EXT = [".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tif", ".tiff"]

# Format output yang didukung
SUPPORTED_FORMATS = ["xlsx", "csv", "json", "txt"]


# =====================================================
# CLI
# =====================================================
def parse_args():
    p = argparse.ArgumentParser("Verifikasi (TFLite saja) -> excel_idpel_kwh.xlsx (2 kolom)")

    p.add_argument("--model", default=MODEL_PATH)
    p.add_argument("--labels", default=LABELS_PATH)
    p.add_argument("--src", default=SRC_DIR)
    p.add_argument("--dst", default=DST_DIR)
    p.add_argument("--log", default=LOG_PATH)
    
    # ✅ Format output
    p.add_argument("--format", choices=SUPPORTED_FORMATS, default="xlsx",
                   help="Format file output (xlsx, csv, json, txt)")
    
    # ✅ Parameter untuk filter gambar
    p.add_argument("--pass-only", action="store_true", default=True,
                   help="Hanya simpan gambar yang lolos verifikasi (default)")
    p.add_argument("--all-images", dest="pass-only", action="store_false",
                   help="Simpan semua gambar")

    p.add_argument("--neg_threshold", type=float, default=NEG_THRESHOLD)
    p.add_argument("--kwh_threshold", type=float, default=KWH_THRESHOLD)
    p.add_argument("--blur_threshold", type=float, default=BLUR_THRESHOLD)
    # Tambah argumen untuk threshold foto gelap
    p.add_argument("--brightness_threshold", type=float, default=BRIGHTNESS_THRESHOLD)
    p.add_argument("--contrast_threshold", type=float, default=CONTRAST_THRESHOLD)

    p.add_argument("--expected_idpel_len", type=int, default=EXPECTED_IDPEL_LEN)

    # ✅ embed thumbnail default ON untuk menampilkan gambar
    p.add_argument("--embed_images", dest="embed_images", action="store_true")
    p.add_argument("--no_embed_images", dest="embed_images", action="store_false")
    p.set_defaults(embed_images=True)

    p.add_argument("--thumb_size", type=int, default=100,
                   help="Ukuran thumbnail dalam pixel")

    return p.parse_args()


# =====================================================
# Labels / Model
# =====================================================
def load_labels(path: str):
    labels = [x.strip() for x in open(path, "r", encoding="utf-8") if x.strip()]
    if len(labels) != 2:
        raise ValueError("labels.txt harus 2 kelas: KWH & NEG")

    low = [l.lower() for l in labels]
    if not any("kwh" in l for l in low):
        raise ValueError("labels.txt harus mengandung label yang ada kata 'kwh'")

    kwh_idx = next(i for i, l in enumerate(low) if "kwh" in l)
    neg_idx = 1 - kwh_idx
    return labels, kwh_idx, neg_idx

def load_interpreter(model_path: str):
    itp = tf.lite.Interpreter(model_path=model_path)
    itp.allocate_tensors()

    in_det = itp.get_input_details()[0]
    out_det = itp.get_output_details()[0]
    h, w = in_det["shape"][1:3]
    dtype = in_det["dtype"]
    return itp, in_det, out_det, (h, w, dtype)

def softmax(x):
    x = np.asarray(x, dtype=np.float32)
    e = np.exp(x - np.max(x))
    return e / (e.sum() + 1e-9)


# =====================================================
# Preprocess (FAST): dari cv2 array, tanpa buka file lagi
# =====================================================
def preprocess_bgr_for_tflite(img_bgr, size_wh, dtype):
    w, h = size_wh
    rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    resized = cv2.resize(rgb, (w, h), interpolation=cv2.INTER_AREA)

    arr = resized.astype(np.float32)
    arr = (arr / 127.5) - 1.0

    if dtype == np.uint8:
        arr = ((arr + 1) * 127.5).clip(0, 255).astype(np.uint8)

    return np.expand_dims(arr, axis=0)


# =====================================================
# IDPEL helpers
# =====================================================
def normalize_idpel_digits(d: str, expected_len: int = 12) -> str:
    d = re.sub(r"\D+", "", str(d or ""))
    if not d:
        return ""
    if expected_len and len(d) > expected_len:
        d = d[-expected_len:]
    return d

def extract_idpel_from_filename(img_path: Path, expected_len: int = 12) -> str:
    stem_digits = re.sub(r"\D+", "", img_path.stem)
    if not stem_digits:
        return ""
    if expected_len and len(stem_digits) > expected_len:
        stem_digits = stem_digits[-expected_len:]
    return stem_digits


# =====================================================
# Quality checkers: Blur detector + Brightness/Contrast detector
# =====================================================
def blur_score(img_bgr) -> float:
    try:
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        return float(cv2.Laplacian(gray, cv2.CV_64F).var())
    except Exception:
        return 0.0

def check_image_quality(img_bgr, brightness_threshold=30.0, contrast_threshold=20.0):
    """
    Check if image is too dark/low contrast
    Returns: (is_good, brightness_score, contrast_score)
    """
    try:
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        
        # Brightness check (rata-rata intensitas pixel)
        brightness = np.mean(gray)
        
        # Contrast check (standar deviasi intensitas pixel)
        contrast = np.std(gray)
        
        is_good = (brightness >= brightness_threshold) and (contrast >= contrast_threshold)
        
        return is_good, float(brightness), float(contrast)
    except Exception:
        return False, 0.0, 0.0


# =====================================================
# Thumbnail untuk Excel
# =====================================================
def create_thumbnail_for_excel(img_bgr, thumb_size=100):
    """Buat thumbnail untuk dimasukkan ke Excel"""
    try:
        h, w = img_bgr.shape[:2]
        
        # Hitung scaling factor
        scale = thumb_size / max(h, w)
        new_w = int(w * scale)
        new_h = int(h * scale)
        
        # Resize image
        resized = cv2.resize(img_bgr, (new_w, new_h), interpolation=cv2.INTER_AREA)
        
        # Convert BGR to RGB
        rgb = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)
        
        # Convert to PIL Image
        from PIL import Image
        pil_img = Image.fromarray(rgb)
        
        # Save to temporary file
        temp_dir = tempfile.gettempdir()
        temp_path = Path(temp_dir) / f"thumb_{int(time.time())}_{hash(str(img_bgr))}.png"
        pil_img.save(str(temp_path), format="PNG", optimize=True)
        
        return temp_path
    except Exception as e:
        print(f"[ERROR] Gagal membuat thumbnail: {e}")
        return None


# =====================================================
# Workbook untuk Excel dengan gambar (HANYA 2 KOLOM)
# =====================================================
def build_workbook_with_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "IDPEL_KWH"

    # ✅ HANYA 2 KOLOM: gambar | idpel
    ws.append(["gambar", "idpel"])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format header untuk 2 kolom
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:B1"

    # Atur lebar kolom
    ws.column_dimensions["A"].width = 20  # Kolom untuk gambar
    ws.column_dimensions["B"].width = 20  # Kolom untuk IDPEL
    
    # Atur tinggi baris untuk gambar
    ws.row_dimensions[1].height = 30  # Header
    for i in range(2, 1000):  # Set default height untuk data rows
        ws.row_dimensions[i].height = 80  # Tinggi untuk thumbnail

    return wb, ws


# =====================================================
# Fungsi untuk menyimpan ke berbagai format
# =====================================================
def save_to_excel_with_images(data, output_path, thumb_size=100):
    """Simpan data ke Excel dengan gambar thumbnail (HANYA 2 KOLOM)"""
    print(f"[INFO] Membuat Excel dengan gambar thumbnail...")
    
    wb, ws = build_workbook_with_images()
    temp_files = []  # Untuk menyimpan path file temporary
    
    try:
        for i, row_data in enumerate(data, 2):  # Mulai dari row 2
            img_path, idpel, img_bgr = row_data
            
            # Tambah IDPEL di kolom B
            ws.cell(row=i, column=2, value=idpel)  # Kolom B: IDPEL
            ws.cell(row=i, column=2).alignment = Alignment(vertical="center", horizontal="center")
            
            # Tambah gambar thumbnail jika tersedia
            if img_bgr is not None:
                try:
                    # Buat thumbnail
                    temp_thumb = create_thumbnail_for_excel(img_bgr, thumb_size)
                    if temp_thumb and temp_thumb.exists():
                        # Tambah gambar ke Excel (kolom A)
                        img = XLImage(str(temp_thumb))
                        img.anchor = f"A{i}"  # Kolom A untuk gambar
                        ws.add_image(img)
                        temp_files.append(temp_thumb)
                except Exception as e:
                    print(f"[WARN] Gagal menambahkan thumbnail: {e}")
            
            # Progress indicator
            if i % 50 == 0:
                print(f"[PROGRESS] {i-1}/{len(data)} gambar diproses...")
        
        # Simpan Excel file
        wb.save(str(output_path))
        print(f"[SUCCESS] Excel berhasil disimpan: {output_path}")
        
    finally:
        # Bersihkan file temporary
        for temp_file in temp_files:
            try:
                if temp_file.exists():
                    temp_file.unlink()
            except:
                pass

def save_to_excel_simple(data, output_path):
    """Simpan data ke Excel tanpa gambar (hanya teks) - HANYA 2 KOLOM"""
    wb = Workbook()
    ws = wb.active
    ws.title = "IDPEL_KWH"
    
    # Header HANYA 2 KOLOM
    ws.append(["gambar", "idpel"])
    
    # Data - HANYA IDPEL (gambar tidak ditampilkan)
    for img_path, idpel, _ in data:
        ws.append(["", idpel])  # Kolom gambar kosong, kolom idpel berisi
    
    # Format header
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Atur lebar kolom
    ws.column_dimensions["A"].width = 20  # Kolom gambar (kosong)
    ws.column_dimensions["B"].width = 20  # Kolom IDPEL
    
    wb.save(str(output_path))
    print(f"[INFO] Excel sederhana disimpan: {output_path}")

def save_to_csv(data, output_path):
    """Simpan data ke format CSV - HANYA 2 KOLOM"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["gambar", "idpel"])  # ✅ HANYA 2 HEADER
        
        for img_path, idpel, _ in data:
            writer.writerow(["", idpel])  # Kolom gambar kosong
    
    print(f"[INFO] CSV disimpan: {output_path}")

def save_to_json(data, output_path):
    """Simpan data ke format JSON - HANYA 2 FIELD"""
    result = []
    for img_path, idpel, _ in data:
        result.append({
            "idpel": idpel  # ✅ HANYA IDPEL
        })
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({"data_kwh": result, "total": len(result)}, f, indent=2, ensure_ascii=False)
    
    print(f"[INFO] JSON disimpan: {output_path}")

def save_to_txt(data, output_path):
    """Simpan data ke format TXT - HANYA 2 KOLOM"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 50 + "\n")
        f.write("DATA IDPEL KWH METER\n")
        f.write("=" * 50 + "\n\n")
        
        f.write(f"{'No.':<5} {'IDPEL':<15}\n")
        f.write("-" * 25 + "\n")
        
        for i, (_, idpel, _) in enumerate(data, 1):
            f.write(f"{i:<5} {idpel:<15}\n")
        
        f.write(f"\nTotal: {len(data)} IDPEL\n")
        f.write(f"Tanggal: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}\n")
    
    print(f"[INFO] TXT disimpan: {output_path}")


# =====================================================
# MAIN
# =====================================================
def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

    args = parse_args()

    # Import PIL di sini
    from PIL import Image

    labels, kwh_idx, neg_idx = load_labels(args.labels)
    itp, in_det, out_det, meta = load_interpreter(args.model)
    in_h, in_w, dtype = meta
    model_size = (in_w, in_h)

    src_dir = Path(args.src)
    dst_dir = Path(args.dst)
    dst_dir.mkdir(parents=True, exist_ok=True)

    # ✅ NAMA FILE OUTPUT FIXED
    log_path = Path(args.log)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    # ✅ Siapkan data
    valid_data = []
    images = [p for p in src_dir.rglob("*") if p.is_file() and p.suffix.lower() in IMG_EXT]
    images.sort(key=lambda x: x.name.lower())

    total = 0
    passed = 0
    failed = 0

    print(f"[INFO] Memulai verifikasi {len(images)} gambar...")
    print(f"[INFO] Output: {log_path.name}")
    print(f"[INFO] Format: {args.format.upper()}")
    print(f"[INFO] Mode: {'Hanya gambar PASS' if args.pass_only else 'Semua gambar'}")
    print(f"[INFO] Gambar di Excel: {'YA' if args.embed_images and args.format == 'xlsx' else 'TIDAK'}")
    print("-" * 60)

    for img_path in images:
        try:
            img_bgr = cv2.imread(str(img_path))
            if img_bgr is None:
                print(f"[ERROR] Gagal membaca: {img_path.name}")
                failed += 1
                continue

            idpel = extract_idpel_from_filename(img_path, args.expected_idpel_len)
            idpel = normalize_idpel_digits(idpel, args.expected_idpel_len)

            # ===== TFLite inference =====
            inp = preprocess_bgr_for_tflite(img_bgr, model_size, dtype)
            itp.set_tensor(in_det["index"], inp)
            itp.invoke()
            raw = np.squeeze(itp.get_tensor(out_det["index"]))

            if np.ndim(raw) == 0:
                probs = np.array([1 - float(raw), float(raw)], dtype=np.float32)
            else:
                probs = softmax(raw.astype(np.float32))

            p_kwh = float(probs[kwh_idx])
            p_neg = float(probs[neg_idx])
            pred = "KWH" if p_kwh >= p_neg else "NEG"

            # ===== Quality checks =====
            status_ok = True
            
            # 1. Check blur
            bscore = blur_score(img_bgr)
            if bscore < args.blur_threshold:
                status_ok = False
                print(f"[FAIL] {img_path.name}: Blur ({bscore:.1f})")

            # 2. Check brightness/contrast
            is_good_quality, brightness, contrast = check_image_quality(
                img_bgr, args.brightness_threshold, args.contrast_threshold)
            if not is_good_quality:
                status_ok = False
                print(f"[FAIL] {img_path.name}: Kualitas rendah")

            # 3. Check pagar
            if "pagar" in img_path.name.lower() or "pagar" in str(img_path.parent).lower():
                status_ok = False
                print(f"[FAIL] {img_path.name}: Mengandung 'pagar'")

            # 4. Check classification
            if pred == "NEG":
                status_ok = False
                print(f"[FAIL] {img_path.name}: NEG ({p_neg:.2f})")
            elif pred == "KWH" and p_kwh < args.kwh_threshold:
                status_ok = False
                print(f"[FAIL] {img_path.name}: KWH rendah ({p_kwh:.2f})")

            total += 1
            
            # Simpan data jika valid
            if status_ok:
                passed += 1
                valid_data.append((img_path, idpel, img_bgr if args.embed_images else None))
                print(f"[PASS] {img_path.name}: IDPEL={idpel}")
            elif not args.pass_only:
                failed += 1
                valid_data.append((img_path, idpel, img_bgr if args.embed_images else None))
                print(f"[FAIL] {img_path.name}: IDPEL={idpel}")
            else:
                failed += 1

            if total % 50 == 0:
                print(f"[PROGRESS] {total}/{len(images)} gambar diproses...")

        except Exception as e:
            print(f"[ERROR] {img_path.name}: {e}")
            failed += 1

    # ✅ SIMPAN DATA
    if valid_data:
        if args.format == "xlsx":
            if args.embed_images:
                save_to_excel_with_images(valid_data, log_path, args.thumb_size)
            else:
                save_to_excel_simple(valid_data, log_path)
        elif args.format == "csv":
            save_to_csv(valid_data, log_path.with_suffix(".csv"))
        elif args.format == "json":
            save_to_json(valid_data, log_path.with_suffix(".json"))
        elif args.format == "txt":
            save_to_txt(valid_data, log_path.with_suffix(".txt"))
    else:
        print(f"[WARN] Tidak ada data valid!")
        # Buat file Excel kosong dengan 2 kolom
        wb = Workbook()
        ws = wb.active
        ws.append(["gambar", "idpel"])
        wb.save(str(log_path))
        print(f"[INFO] File Excel kosong dibuat: {log_path}")

    # SUMMARY
    print("\n" + "=" * 60)
    print("HASIL AKHIR")
    print("=" * 60)
    print(f"Total dipindai    : {total}")
    print(f"Valid (PASS)      : {passed}")
    print(f"Tidak valid       : {failed}")
    
    if total > 0:
        success_rate = (passed/total*100)
        print(f"Tingkat sukses    : {success_rate:.1f}%")
    
    print(f"File output       : {log_path.name}")
    print(f"Kolom Excel       : gambar | idpel")
    print(f"Jumlah data       : {len(valid_data)}")
    print("=" * 60)

if __name__ == "__main__":
    main()